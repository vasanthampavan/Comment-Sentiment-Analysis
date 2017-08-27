from bs4 import BeautifulSoup
import urllib.request
import json
import openpyxl
from openpyxl.styles import Font
from commentHelper import *

'''Script to retrieve Kinja comments into an excel file'''

# row to start inputting data (change if needed)
excelRow = 2
wb = openpyxl.load_workbook('KinjaComments.xlsx')
sheet = wb.get_sheet_by_name("Sheet1")

#Change index here to look for specific articles
debugCounter = 1

articleStartIndex = 0
articleEndIndex = 32

print("Note: This program only works on Kinja websites, files/links intended to be scraped should include a 10 digit article code.")
print("")
# userInput = input("Press Enter for mass scraping OR paste link/article code for specific link scraping: ")
# if userInput == "":
#     getSpecific = False

# articleStartIndex = int(input("Please choose the start index for the articles you want to scrape: ")) - 1
# articleEndIndex = articleStartIndex + int(input("Please choose how many articles you would like to scrape: "))

# else:
#     getSpecific = True

articleCodes = getSplinterCodes()

for articleIndex in range(articleStartIndex, articleEndIndex):
#Index to keep track of the comments (used to change link and get new comments)
    startIndex = 0
    numberOfComments = 0
    approvedChildComments = 0

    currentCode = articleCodes[articleIndex]
    webURL = "http://splinternews.com/{}".format(currentCode)

    # if getSpecific:
    #     webURL = userInput
    #     articleCodes[articleIndex] = findCode(webURL)

    #Open request to webpage
    try:
        web = urllib.request.urlopen(webURL)
    except:
        print("Error, cannot open URL: " + webURL)
        break

    soup = BeautifulSoup(web.read(), "html.parser")

    #Find the specific HTML element that holds the number of total replies
    try:
        headline = findHeadline(soup)
        totalNumComments = findReplies(soup)
    except:
        print("Error, cannot find headline or totalcomments")

    headlineRow = excelRow
    sheet.cell(row=headlineRow, column=1).value = currentCode
    sheet.cell(row=headlineRow, column=2).hyperlink = webURL
    sheet.cell(row=headlineRow+1, column=2).value = "Article Title: " + headline

    #stats to keep track of
    avgMainWord = 0
    avgMainChar = 0
    avgChildWord = 0
    avgChildChar = 0
    imageCount = 0

    #Keep looping until we get all comments (calling different JSON links)
    while startIndex < totalNumComments:

        #Link can be changed to included non approved comments as well
        jsonURL = "http://splinternews.com/api/comments/views/replies/{0}?dap=true&startIndex={1}&maxReturned" \
                  "=100&maxChildren=100&approvedOnly=true&cache=true".format(articleCodes[articleIndex], startIndex)


        page = urllib.request.urlopen(jsonURL).read()
        pageString = page.decode('utf-8')
        decoded = json.loads(pageString)
        dataSet = decoded["data"]["items"]

        counter = 0
        while counter < len(dataSet):

            mainComment = dataSet[counter]["reply"]["deprecatedFullPlainText"]

            try:
                imageColumn = 6
                imageSet = dataSet[counter]["reply"]["images"]
                imageCounter = 0
                imageLink = ""
                while imageCounter < len(imageSet):
                    imageType = imageSet[imageCounter]["format"]
                    imageId = imageSet[imageCounter]["id"]
                    imageLink = "https://i.kinja-img.com/gawker-media/image/upload/"+str(imageId)+"."+str(imageType)
                    sheet.cell(row = excelRow, column = imageColumn).hyperlink = str(imageLink)
                    imageCounter += 1
                    imageColumn += 1
                    imageCount += 1
            except:
                print("no main comment image")


            #making sure the comment is not empty
            if mainComment != "":
                mainCommentWordCount = countWords(mainComment)
                mainCommentCharacterCount = countCharacters(mainComment)
                avgMainWord += mainCommentWordCount
                avgMainChar += mainCommentCharacterCount
                try:
                    if mainComment.strip() != "":
                        sheet.cell(row=excelRow, column=3).value = mainComment
                    else:
                        sheet.cell(row=excelRow, column=3).value = "(main image comment)"
                except:
                    sheet.cell(row=excelRow, column=3).value = "error string"
                sheet.cell(row = excelRow, column=3).font = Font(bold = True)
                sheet.cell(row = excelRow, column=4).value = mainCommentWordCount
                sheet.cell(row = excelRow, column=5).value = mainCommentCharacterCount
                excelRow+=1
            elif imageLink != None:
                sheet.cell(row = excelRow, column = 3).value = "(main image comment)"
                sheet.cell(row = excelRow, column = 3).font = Font(bold = True)
                sheet.cell(row = excelRow, column = 4).value = 0
                sheet.cell(row = excelRow, column = 5).value = 0
                excelRow+=1

            numberOfComments+=1

            childSet = dataSet[counter]["children"]["items"]

            childCounter = 0
            while childCounter < len(childSet):

                childComment = childSet[childCounter]["deprecatedFullPlainText"]

                try:
                    imageSet = childSet[childCounter]["images"]
                    imageColumn = 6
                    imageCounter = 0
                    imageLink = ""
                    while imageCounter < len(imageSet):
                        imageType = imageSet[imageCounter]["format"]
                        imageId = imageSet[imageCounter]["id"]
                        imageLink += "https://i.kinja-img.com/gawker-media/image/upload/"+str(imageId)+"."+str(imageType) + "   "
                        sheet.cell(row = excelRow, column = imageColumn).hyperlink = str(imageLink)
                        imageCounter += 1
                        imageColumn += 1
                        imageCount += 1
                except:
                    imageLink = "nil"
                    print("no child comment image")

                if childComment != "":
                    childWordLen = countWords(childComment)
                    childCharLen = countCharacters(childComment)
                    avgChildWord += childWordLen
                    avgChildChar += childCharLen

                    try:
                        if childComment.strip() != "":
                            sheet.cell(row=excelRow, column=3).value = childComment
                        else:
                            sheet.cell(row=excelRow, column=3).value = "(child image comment)"
                    except:
                        sheet.cell(row=excelRow, column=3).value = "error string"
                    sheet.cell(row = excelRow, column = 4).value = childWordLen
                    sheet.cell(row = excelRow, column = 5).value = childCharLen
                    excelRow+=1
                    approvedChildComments+= 1
                elif imageLink != None:
                    sheet.cell(row = excelRow, column = 3).value = "(child image comment)"
                    sheet.cell(row = excelRow, column = 4).value = 0
                    sheet.cell(row = excelRow, column = 5).value = 0
                    excelRow+=1
                    approvedChildComments+= 1
                childCounter+=1
            counter += 1
        startIndex+=100
        print("Adding another 100 to get comments " + str(startIndex))



    #code to parse into excel document chosen above
    sheet.cell(row = headlineRow+2, column = 2).value = "Number of total comments: {}".format(str(totalNumComments))
    sheet.cell(row = headlineRow+3, column = 2).value = "Number of total posted comments: {}".\
        format(str(numberOfComments + approvedChildComments))
    sheet.cell(row = headlineRow+4, column = 2).value = "Number of main comments: {}"\
        .format(str(numberOfComments))
    sheet.cell(row = headlineRow+5, column = 2).value = "Number of child comments: {}".\
        format(str(approvedChildComments))

    try:
        sheet.cell(row = headlineRow+6, column = 2).value = "Average Main Comment Word Count: {}".\
                format(str(avgMainWord/numberOfComments))
        sheet.cell(row = headlineRow+7, column = 2).value = "Average Main Comment Character Count: {}".\
                    format(str(avgMainChar/numberOfComments))
    except:
        sheet.cell(row = headlineRow+6, column = 2).value = "Average Main Comment Word Count: {}".\
                    format("0")
        sheet.cell(row = headlineRow+7, column = 2).value = "Average Main Comment Character Count: {}".\
                format("0")

    try:
        sheet.cell(row = headlineRow+8, column = 2).value = "Average Child Comment Word Count: {}".\
                    format(str(avgChildWord/approvedChildComments))
        sheet.cell(row = headlineRow+9, column = 2).value = "Average Child Comment Character Count: {}".\
                format(str(avgChildChar/approvedChildComments))
    except:
        sheet.cell(row = headlineRow+8, column = 2).value = "Average Child Comment Word Count: {}".\
                    format("0")
        sheet.cell(row = headlineRow+9, column = 2).value = "Average Child Comment Character Count: {}".\
                format("0")

    sheet.cell(row = headlineRow+10, column = 2).value = "Number of images: {}".\
            format(str(imageCount))


    print("Article {} done".format(debugCounter))
    debugCounter+=1

    if(numberOfComments < 9):
        excelRow += 12
    else:
        excelRow += 2
    wb.save('KinjaComments.xlsx')

wb.save('KinjaComments.xlsx')
